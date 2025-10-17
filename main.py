#!/usr/bin/env python3
"""
Compare two Excel/CSV files (one in ./before, one in ./after) whose date columns (e.g., 10/19, 10/26, …) are aligned.
Outputs diff_report.xlsx with:
  - value_changes: row-by-row per-date diffs (after - before) for matching keys
  - added_rows: rows only in AFTER
  - removed_rows: rows only in BEFORE
  - summary: totals per date, plus counts

Assumptions
- Each folder (./before and ./after) contains exactly one data file (.xlsx/.xls/.csv).
- Non-date columns are key columns used to match rows (auto-detected).
- Date columns have headers like M/D or M/D/YY(YY). (You can adjust the regex DATE_COL_RE if needed.)
"""

from __future__ import annotations
import sys
import re
from pathlib import Path
from typing import List, Tuple

import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import streamlit as st
from io import BytesIO

# Global constants
DATE_COL_RE = re.compile(r"^\s*\d{1,2}/\d{1,2}(?:/\d{2,4})?\s*$")  # e.g., 10/19, 11/2, 11/16/2024
POSITIVE_THRESHOLD = 100  # Threshold for positive delta highlighting
NEGATIVE_THRESHOLD = -100  # Threshold for negative delta highlighting (use negative value)


def find_single_file(folder: Path) -> Path:
    if not folder.exists() or not folder.is_dir():
        sys.exit(f"Folder not found: {folder}")
    candidates = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {'.xlsx', '.xls', '.csv'}]
    if len(candidates) == 0:
        sys.exit(f"No .xlsx/.xls/.csv file found in {folder}")
    if len(candidates) > 1:
        sys.exit(f"Expected exactly one file in {folder}, found {len(candidates)}: {[p.name for p in candidates]}")
    return candidates[0]


def _normalize_table(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize headers, trim strings, and drop empty columns."""

    df.columns = [str(c).strip() for c in df.columns]
    df = df.map(lambda x: "" if pd.isna(x) else str(x).strip())
    return df.loc[:, ~(df == "").all(axis=0)]


def load_table_from_path(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        df = pd.read_excel(path, dtype=str, engine="openpyxl")

    return _normalize_table(df)


def detect_date_and_key_cols(df: pd.DataFrame) -> Tuple[List[str], List[str]]:
    """
    Identify date columns based on either:
      - M/D(/YY) header strings (regex), OR
      - Headers that parse as dates (e.g., Excel date headers become '2024-10-19 00:00:00').
    Keeps the original column order.
    """
    date_cols = []
    for c in df.columns:
        cs = str(c).strip()
        if DATE_COL_RE.match(cs):
            date_cols.append(c)
            continue
        try:
            parsed = pd.to_datetime(cs, errors="raise")
            if not pd.isna(parsed):
                date_cols.append(c)
        except Exception:
            pass

    key_cols = [c for c in df.columns if c not in date_cols]
    if not key_cols:
        raise SystemExit("Could not determine key columns (everything looks like a date). Adjust DATE_COL_RE.")
    if not date_cols:
        raise SystemExit("Could not detect any date columns. Check your headers or DATE_COL_RE.")
    return key_cols, date_cols


def coerce_numeric_dates(df: pd.DataFrame, date_cols: List[str]) -> pd.DataFrame:
    """
    Convert date columns to numeric:
    - Remove commas/spaces
    - Convert parentheses negatives: (123) -> -123
    - Treat blanks and dashes as 0
    """
    out = df.copy()
    for c in date_cols:
        s = out[c].astype(str).str.strip()

        s = (
            s.str.replace(r"\s+", "", regex=True)
            .str.replace(",", "", regex=False)
            .str.replace(r"^\((.*)\)$", r"-\1", regex=True)
        )
        s = s.replace({"": "0", "-": "0", "—": "0", "–": "0"})

        out[c] = pd.to_numeric(s, errors="coerce").fillna(0)
    return out


def make_key_index(df: pd.DataFrame, key_cols: List[str]) -> pd.Index:
    cleaned = df[key_cols].fillna("").map(lambda x: str(x).strip())
    tuples = list(map(tuple, cleaned.values.tolist()))

    if len(key_cols) > 1:
        return pd.MultiIndex.from_tuples(tuples, names=key_cols)
    else:
        return pd.Index([t[0] for t in tuples], name=key_cols[0])


def format_value_changes_stacked(key_cols: List[str], date_cols: List[str],
                                 dfb_shared: pd.DataFrame, dfa_shared: pd.DataFrame) -> pd.DataFrame:
    """
    Format value_changes with stacked rows for each key: delta, before, after
    """
    rows = []
    for idx_pos in range(len(dfb_shared)):
        key_idx = dfb_shared.index[idx_pos]
        if isinstance(key_idx, tuple):
            key_vals = key_idx
        else:
            key_vals = (key_idx,)

        before_row = dfb_shared.iloc[idx_pos]
        after_row = dfa_shared.iloc[idx_pos]
        delta_row = after_row - before_row

        if not (delta_row != 0).any():
            continue

        key_dict = {f"key_{i + 1}": key_vals[i] for i in range(len(key_vals))}

        # Delta row
        row_delta = {**key_dict, "type": "delta"}
        for dc in date_cols:
            delta_val = delta_row[dc]
            row_delta[dc] = delta_val if delta_val != 0 else ""
        rows.append(row_delta)

        # Before row
        row_before = {**key_dict, "type": "before"}
        for dc in date_cols:
            delta_val = delta_row[dc]
            if delta_val != 0:
                row_before[dc] = before_row[dc]
            else:
                row_before[dc] = ""
        rows.append(row_before)

        # After row
        row_after = {**key_dict, "type": "after"}
        for dc in date_cols:
            delta_val = delta_row[dc]
            if delta_val != 0:
                row_after[dc] = after_row[dc]
            else:
                row_after[dc] = ""
        rows.append(row_after)

    if not rows:
        key_cols_out = [f"key_{i + 1}" for i in range(len(key_cols))]
        return pd.DataFrame(columns=key_cols_out + ["type"] + date_cols)

    result = pd.DataFrame(rows)
    key_cols_out = [f"key_{i + 1}" for i in range(len(key_cols))]
    cols_order = key_cols_out + ["type"] + date_cols
    return result[cols_order]


def prepare_diff_results(df_before: pd.DataFrame, df_after: pd.DataFrame):
    """Prepare comparison DataFrames and metadata for report generation."""

    warnings = []
    key_cols_b, date_cols_b = detect_date_and_key_cols(df_before)
    key_cols_a, date_cols_a = detect_date_and_key_cols(df_after)

    if date_cols_b != date_cols_a:
        if set(date_cols_b) != set(date_cols_a):
            raise SystemExit(
                f"Date columns differ.\nBefore: {date_cols_b}\nAfter : {date_cols_a}"
            )
        df_after = df_after[key_cols_a + date_cols_b]
        date_cols_a = date_cols_b

    if key_cols_b != key_cols_a:
        shared_keys = [c for c in key_cols_b if c in key_cols_a]
        if not shared_keys:
            raise SystemExit("No shared key columns found between BEFORE and AFTER.")
        warnings.append(
            "Warning: key column names differ between files. Using shared columns for alignment."
        )
        df_before_keys = shared_keys
        df_after_keys = shared_keys
    else:
        df_before_keys = key_cols_b
        df_after_keys = key_cols_a

    df_before_num = coerce_numeric_dates(df_before, date_cols_b)
    df_after_num = coerce_numeric_dates(df_after, date_cols_a)

    idx_before = make_key_index(df_before_num, df_before_keys)
    idx_after = make_key_index(df_after_num, df_after_keys)

    agg_funcs = {c: "sum" for c in date_cols_b}
    dfb = df_before_num.groupby(idx_before, dropna=False).agg({**{k: "first" for k in df_before_keys}, **agg_funcs})
    dfa = df_after_num.groupby(idx_after, dropna=False).agg({**{k: "first" for k in df_after_keys}, **agg_funcs})

    keys_before = list(dict.fromkeys(idx_before))
    keys_after = list(dict.fromkeys(idx_after))

    dfb_index_set = set(dfb.index)
    dfa_index_set = set(dfa.index)

    added_keys = [k for k in keys_after if k not in dfb_index_set]
    removed_keys = [k for k in keys_before if k not in dfa_index_set]
    shared_keys = [k for k in keys_before if k in dfa_index_set]

    added_rows = (
        dfa.loc[added_keys].reset_index()
        if added_keys
        else pd.DataFrame(columns=["__key__"] + df_after.columns.tolist())
    )
    removed_rows = (
        dfb.loc[removed_keys].reset_index()
        if removed_keys
        else pd.DataFrame(columns=["__key__"] + df_before.columns.tolist())
    )

    if shared_keys:
        dfb_shared = dfb.loc[shared_keys, date_cols_b]
        dfa_shared = dfa.loc[shared_keys, date_cols_b]
    else:
        dfb_shared = pd.DataFrame(columns=date_cols_b)
        dfa_shared = pd.DataFrame(columns=date_cols_b)

    value_changes = format_value_changes_stacked(df_before_keys, date_cols_b, dfb_shared, dfa_shared)

    if not dfb_shared.empty:
        total_before = dfb_shared.sum().rename("total_before")
        total_after = dfa_shared.sum().rename("total_after")
        total_delta = (total_after - total_before).rename("total_delta")
        summary = pd.concat([total_before, total_after, total_delta], axis=1).reset_index().rename(
            columns={"index": "date"}
        )
    else:
        summary = pd.DataFrame(columns=["date", "total_before", "total_after", "total_delta"])

    return {
        "value_changes": value_changes,
        "added_rows": added_rows.reset_index(drop=True),
        "removed_rows": removed_rows.reset_index(drop=True),
        "summary": summary,
        "date_cols": date_cols_b,
        "num_key_cols": len(df_before_keys),
        "warnings": warnings,
    }


def write_diff_report(results, output):
    """Write the comparison report to a file path or file-like buffer."""

    with pd.ExcelWriter(output, engine="openpyxl") as xw:
        results["value_changes"].to_excel(xw, index=False, sheet_name="value_changes")
        results["added_rows"].to_excel(xw, index=False, sheet_name="added_rows")
        results["removed_rows"].to_excel(xw, index=False, sheet_name="removed_rows")
        results["summary"].to_excel(xw, index=False, sheet_name="summary")

        ws = xw.sheets["value_changes"]
        apply_formatting_to_value_changes(ws, results["date_cols"], results["num_key_cols"])

        row_idx = 2
        max_row = ws.max_row
        num_key_cols = results["num_key_cols"]
        while row_idx <= max_row:
            row_type = ws.cell(row=row_idx, column=num_key_cols + 1).value
            if row_type == "delta":
                before_idx = row_idx + 1
                after_idx = row_idx + 2
                if after_idx <= max_row:
                    ws.row_dimensions.group(before_idx, after_idx, outline_level=1)
                    ws.row_dimensions[before_idx].hidden = True
                    ws.row_dimensions[after_idx].hidden = True
                row_idx = after_idx + 1
            else:
                row_idx += 1


def format_date_header(col_name):
    """Convert date column header to simple date format"""
    try:
        parsed = pd.to_datetime(col_name)
        return parsed.strftime("%m/%d")
    except:
        return col_name


def apply_formatting_to_value_changes(ws, date_cols, num_key_cols):
    """Apply formatting to the value_changes worksheet"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format date headers
    for col_idx, date_col in enumerate(date_cols, start=num_key_cols + 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = format_date_header(str(date_col))
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply borders and formatting to all cells
    for row_idx in range(2, ws.max_row + 1):
        row_type = ws.cell(row=row_idx, column=num_key_cols + 1).value

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            # Format date value cells
            if col_idx > num_key_cols + 1 and cell.value != "":
                try:
                    val = float(cell.value) if isinstance(cell.value, (int, float)) else float(str(cell.value))
                    cell.value = val
                    cell.number_format = '0'

                    if row_type == "delta":
                        # Light yellow background for delta rows by default
                        cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

                        # Conditional formatting for delta values (overrides yellow)
                        if val <= NEGATIVE_THRESHOLD:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006", bold=True)
                        elif val >= POSITIVE_THRESHOLD:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100", bold=True)
                except (ValueError, TypeError):
                    pass

            # Light yellow for delta rows (non-value cells)
            if row_type == "delta" and col_idx <= num_key_cols + 1:
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")


def main(base_dir: str = "."):
    base = Path(base_dir).resolve()
    before_file = find_single_file(base / "before")
    after_file = find_single_file(base / "after")

    print(f"BEFORE: {before_file}")
    print(f"AFTER : {after_file}")

    df_before = load_table_from_path(before_file)
    df_after = load_table_from_path(after_file)

    results = prepare_diff_results(df_before, df_after)

    for warning in results.get("warnings", []):
        print(warning)

    out_path = base / "diff_report.xlsx"
    write_diff_report(results, out_path)

    print(f"Wrote: {out_path}")


def load_table_from_upload(file_bytes, filename):
    """Load table from uploaded file."""

    if filename.lower().endswith(".csv"):
        df = pd.read_csv(file_bytes, dtype=str)
    else:
        df = pd.read_excel(file_bytes, dtype=str, engine="openpyxl")

    return _normalize_table(df)


st.set_page_config(page_title="Excel Diff Tool", layout="wide")
st.title("📊 Excel/CSV Diff Comparison Tool")
st.write("Upload two files (before & after) to generate a detailed comparison report.")

col1, col2 = st.columns(2)

with col1:
    st.subheader("Before File")
    before_file = st.file_uploader("Upload BEFORE file", type=['xlsx', 'xls', 'csv'], key='before')

with col2:
    st.subheader("After File")
    after_file = st.file_uploader("Upload AFTER file", type=['xlsx', 'xls', 'csv'], key='after')

if "report_bytes" not in st.session_state:
    st.session_state.report_bytes = None
    st.session_state.summary_df = None
    st.session_state.uploaded_pair = None
    st.session_state.report_warnings = []


if before_file and after_file:
    try:
        current_pair = (before_file.name, after_file.name)
        if st.session_state.uploaded_pair != current_pair:
            st.session_state.report_bytes = None
            st.session_state.summary_df = None
            st.session_state.uploaded_pair = current_pair
            st.session_state.report_warnings = []

        st.info("Processing files...")

        df_before = load_table_from_upload(before_file, before_file.name)
        df_after = load_table_from_upload(after_file, after_file.name)

        st.success("✓ Files loaded successfully")

        with st.expander("Preview: Before File"):
            st.dataframe(df_before.head(), width="stretch")

        with st.expander("Preview: After File"):
            st.dataframe(df_after.head(), width="stretch")

        if st.button("🔍 Generate Comparison Report"):
            try:
                results = prepare_diff_results(df_before, df_after)
                report_buffer = BytesIO()
                write_diff_report(results, report_buffer)
                report_buffer.seek(0)

                st.session_state.report_bytes = report_buffer.getvalue()
                st.session_state.summary_df = results["summary"]
                st.session_state.report_warnings = results.get("warnings", [])

            except SystemExit as exc:
                st.error(str(exc))
            except Exception as exc:  # pragma: no cover - safety net for unexpected errors
                st.error(f"Error generating report: {exc}")

        if st.session_state.report_bytes:
            for warning in st.session_state.report_warnings:
                st.warning(warning)

            st.success("Report generated! Click below to download:")
            st.download_button(
                label="⬇️ Download Comparison Report",
                data=st.session_state.report_bytes,
                file_name="diff_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            if st.session_state.summary_df is not None and not st.session_state.summary_df.empty:
                st.subheader("Summary Totals")
                st.dataframe(st.session_state.summary_df, width="stretch")

    except Exception as e:
        st.error(f"Error: {str(e)}")
else:
    st.warning("Please upload both files to proceed")

if __name__ == "__main__":
    base_dir = sys.argv[1] if len(sys.argv) > 1 else "."
    main(base_dir)